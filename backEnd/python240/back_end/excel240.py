import sys
import time
from openpyxl import load_workbook, Workbook
from pathlib import Path
from openpyxl.worksheet.datavalidation import DataValidation
import validateData
from flask import Flask,render_template,jsonify
from flask_restful import Resource, Api
from flask_cors import CORS


app = Flask(__name__)
CORS(app)

# https://openpyxl.readthedocs.io/en/stable/validation.html
wb = load_workbook(filename = 'excel230.xlsx')
#this is the same as get_sheet_by_name in python3
cform = wb['customers']
pform =  wb['products']
iform = wb['invoices']
lform = wb['invoice_line']


# print(customers.title)
# print(customers.max_row)
# print(customers.max_column)


# this is using xlrd:
# from xlrd import open_workbook
# book = open_workbook("excel230.xlsx")
# sheet = book.sheet_by_index(1) #If your data is on sheet 1

# clients=[]

# for row in range(1, sheet.nrows): #start from 1, to leave out row 0
#     client = {'id': sheet.cell(row,0).value, 'name': sheet.cell(row, 1)}
#     # print(sheet.cell(row,0),sheet.cell(row, 1))
#     clients.append(client)

# print(clients)


# this is using openpyxl with validations:

def build_clients(worksheet):

    clients = []

    def client_validates(client):

        result = validateData.validate_natural(client['id']) and \
                 validateData.validate_string(client['name'])
        return result

    passed = 0
    failed = 0

    for row in worksheet.iter_rows():
        client = {'id': row[0].value, 'name': row[1].value}
        if client_validates(client):
            passed +=1
            clients.append(client)
        else:
            failed +=1

    # print(f'   Clients:  imported {passed}, ignored {failed-1}')
    # print(clients)
    return clients
    print('clients', clients)

clients= build_clients(cform)


def build_products(worksheet):

    products = []

    def product_validates(product):

        result = validateData.validate_natural(product['id']) and \
                 validateData.validate_string(product['product']) and\
                 validateData.validate_natural(product['price'])
        return result

    passed = 0
    failed = 0

    for row in worksheet.iter_rows():
        product = {'id': row[0].value, 'product': row[1].value, 'price': row[2].value}
        # print(product)
        if product_validates(product):
            passed +=1
            products.append(product)
        else:
            failed +=1

    # print(f'   Products:  imported {passed}, ignored {failed-1}')
    # print(products)
    return products

products=build_products(pform)


def build_invoices(worksheet):

    invoices = []

    def invoice_validates(invoice):

        result = validateData.validate_natural(invoice['id']) and \
                 validateData.validate_natural(invoice['clientid']) and\
                 validateData.validate_date(invoice['date'])
        return result

    passed = 0
    failed = 0

    for row in worksheet.iter_rows():
        invoice = {'id': row[0].value, 'clientid': row[1].value, 'date': row[2].value}
        # print(invoice)
        if invoice_validates(invoice):
            passed +=1
            invoices.append(invoice)
        else:
            print(invoice) #I have 2 empty lines at the end of the invoices sheet
            failed +=1

    # print(invoices)

    # print(f'   invoices:  imported {passed}, ignored {failed-1}')
    # print(invoices)
    return invoices

invoices = build_invoices(iform)

def build_lines(worksheet):

    lines = []

    def line_validates(line):

        result = validateData.validate_natural(line['id']) and \
                 validateData.validate_natural(line['invoice']) and \
                 validateData.validate_natural(line['item'])
                 
        return result
    
    imported = 0
    ignored = 0

    for row in worksheet.iter_rows():
        line = {'id': row[0].value, 'invoice': row[1].value, 'item': row[2].value}
        # print(line)
        if line_validates(line):
            imported +=1
            lines.append(line)
        else:
            ignored +=1
    
    # print(f'   Lines:    imported {imported}, ignored {ignored-1}')
    # print(lines)
    return lines    

lines= build_lines(lform)


def cross_validate(clients, products, lines, invoices):
    
    test_list = []

    #  clients id is unique:
    print('Data set validation:')
    # using naive method  
    # # to check all unique list elements     
    # for i in range(len(test_list)): 
    #         for i1 in range(len(test_list)): 
    #             if i != i1: 
    #                 if test_list[i] == test_list[i1]: 
    #                     flag = 1

    # return result = True

    # clients: all IDs unique
    # https://www.geeksforgeeks.org/python-check-if-list-contains-all-unique-elements/
    result = True

    if result:
        test_name = '  ' + 'Unique client IDs:'.ljust(29)
        my_set = set()
        for element in clients:
            if element['id'] in my_set:
                offending_value = element['id']
                print(f"failed. Offending value: {offending_value}")
                result = False
                break
            else:
                my_set.add(element['id'])

        if result:
            print(test_name,'passed')

            result = True


 # products: all IDs unique
    if result:
        test_name = '  ' + 'Unique product IDs:'.ljust(29)
        my_set = set()
        for element in products:
            if element['id'] in my_set:
                offending_value = element['id']
                print(f"failed. Offending value: {offending_value}")
                result = False
                break
            else:
                my_set.add(element['id'])

        if result:
            print(test_name,'passed')

 # lines: all IDs unique
    if result:
        test_name = '  ' + 'Unique line IDs:'.ljust(29)
        my_set = set()
        for element in lines:
            if element['id'] in my_set:
                offending_value = element['id']
                print(f"failed. Offending value: {offending_value}")
                result = False
                break
            else:
                my_set.add(element['id'])

        if result:
            print(test_name,'passed')

    # invoices: all IDs unique
    if result:
        test_name = '  ' + 'Unique invoice IDs:'.ljust(29)
        my_set = set()
        for element in invoices:
            if element['id'] in my_set:
                offending_value = element['id']
                print(f"failed. Offending value: {offending_value}")
                result = False
                break
            else:
                my_set.add(element['id'])

        if result:
            print(test_name,'passed')

   # lines: all invoice IDs valid
    if result:
        test_name = '  ' + 'Valid invoice IDs in invoice-item lines:'.ljust(29)        
        my_set = set()
        for element in invoices:
            my_set.add(element['id'])
        
        for element in lines:
            if element['invoice'] not in my_set:
                offending_value = element['invoice']
                print(test_name, f"failed. Offending value: {offending_value}")
                result = False
                break
        
        if result:
            print(test_name,'passed')

    # lines: all product IDs valid
    if result:
        test_name = '  ' + 'Valid product IDs in invoice-item lines:'.ljust(29)
        my_set = set()
        for element in products:
            my_set.add(element['id'])
        
        for element in lines:
            if element['item'] not in my_set:
                offending_value = element['item']
                print(test_name, f"failed. Offending value: {offending_value}")
                result = False
                break
        
        if result:
            print(test_name,'passed')

    # invoices: all client IDs valid
    if result:
        test_name = '  ' + 'Valid client IDs in invoices:'.ljust(29)
        my_set = set()
        for element in clients:
            my_set.add(element['id'])
        
        for element in invoices:
            if element['clientid'] not in my_set:
                offending_value = element['clientid']
                print(test_name, f"failed. Offending value: {offending_value}")
                result = False
                break
        
        if result:
            print(test_name,'passed')


    if result:
        print('Data set validation complete.\n')
    else:
        print('Data set validation failed.\n')

    return result

cross_validate(clients, products, lines, invoices)

def find_key(data_table, data_key, data_value):

    result = next((element for element in data_table if element[data_key] == data_value), None)
    return result

def print_invoice (invoice, clients, products, lines, invoices):

    # Formatting assumptions:
    # SKU 4 chars, item name 20 chars, price 5 chars, quantity 3 chars, subtotal 8
    print () #create a empty line
    print ('Invoice # ',invoice['id'])
    print ('Client:   ',find_key(clients, 'id', invoice['clientid'])['name'])
    print ('Date:     ',invoice['date'].strftime('%Y-%m-%d'))
    print ('SKU'.ljust(10),'Item'.ljust(21),'Price'.ljust(6),'Qty'.ljust(4),'Subtotal'.ljust(9))
    print ('-'*53)
    sales = {}
    for element in lines:
        if (element['invoice'] == invoice['id']):    
            sold_item = element['item']
            sold_item_count = sales.get(sold_item,0) + 1
            sales[sold_item] = sold_item_count

    total = 0
    for element in sales:
        item = find_key(products, 'id', element)
        item_sku = str(element).ljust(10)
        item_name = item['product'].ljust(21)
        item_price = str(item['price']).ljust(6)
        item_qty = str(sales[element]).ljust(4)
        subtotal = item['price'] * sales[element]
        item_subtotal = str(subtotal).ljust(9)
        print(item_sku, item_name, item_price, item_qty, item_subtotal)
        total += subtotal

    print ('-'*53)
    print ('Total: '.ljust(44), total)
    print ()

def main(argv):

    print('Arguments:', argv)
    return_code = 0
    db_name = 'excel230.xlsx'
    print('Using default db:',db_name)
    while True:
        invoice_number = input('Please enter invoice number to generate or "quit" to finish: ')
        if invoice_number == "quit":
            break
        elif invoice_number.isdigit() and int(invoice_number) > 0:
            invoice_number = int(invoice_number)
            invoice = next((element for element in invoices if element['id'] == invoice_number), None)
            if invoice is None:
                print(f'Invoice not in dataset: {invoice_number}')
            else:
                print_invoice(invoice, clients, products, lines, invoices)
        else:
            print(f'Invalid value for invoice number: {invoice_number}')

    return return_code

# if __name__ == '__main__':
#     main(sys.argv)

# an validate ecustomers, products, lines, invoicesxample from openpyxl website:
# # Create the workbook and worksheet we'll be working with
# wb = Workbook()
# ws = wb.active

# # Create a data-validation object with list validation
# dv = DataValidation(type="list", formula1='"Dog,Cat,Bat"', allow_blank=True)

#  # Optionally set a custom error message
# dv.error ='Your entry is not in the list'
# dv.errorTitle = 'Invalid Entry'
# # Optionally set a custom prompt message
# dv.prompt = 'Please select from the list'
# dv.promptTitle = 'List Selection'

# # Add the data-validation object to the worksheet
# ws.add_data_validation(dv)

# # Create some cells, and add them to the data-validation object
# c1 = ws["A1"]
# c1.value = "Dog1"
# dv.add(c1)
# c2 = ws["A2"]
# c2.value = "An invalid value"
# dv.add(c2)

# # Or, apply the validation to a range of cells
# dv.add('B1:B1048576') # This is the same as for the whole of column B

# # Check with a cell is in the validator
# print("B4" in dv)
# print('Dog' in dv)
# # True


@app.route('/clients')
def client_list():
    return render_template("client.html",clients=clients)

@app.route('/client_json')
def client_js_react():
    data = build_clients(cform)
    return jsonify(data)

if __name__ == '__main__':
   app.run(debug = True)

